from openpyxl import Workbook, load_workbook


HEADERS = [
    "data_hora",
    "nome",
    "email",
    "telefone",
    "cep",
    "logradouro",
    "bairro",
    "cidade",
    "uf",
    "numero",
    "complemento",
]


class ExcelRepo:
    def __init__(self, file_path: str = "cadastros.xlsx", sheet_name: str = "usuarios"):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def _ensure_workbook(self):
        try:
            wb = load_workbook(self.file_path)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
                # garante cabeçalho na linha 1
                if ws.max_row < 1 or ws["A1"].value != HEADERS[0]:
                    if ws.max_row == 0:
                        ws.append(HEADERS)
                    else:
                        for i, h in enumerate(HEADERS, start=1):
                            ws.cell(row=1, column=i, value=h)
            else:
                ws = wb.create_sheet(self.sheet_name)
                ws.append(HEADERS)

            # remove sheet padrão se existir e estiver vazio
            if "Sheet" in wb.sheetnames:
                ws_default = wb["Sheet"]
                if ws_default.max_row == 1 and ws_default.max_column == 1 and ws_default["A1"].value is None:
                    wb.remove(ws_default)

            wb.save(self.file_path)

        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(HEADERS)
            wb.save(self.file_path)

    def append_user(self, row: list):
        self._ensure_workbook()
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]
        ws.append(row)
        wb.save(self.file_path)

    def list_users(self) -> list[dict]:
        """
        Retorna lista de dicts com uma chave extra:
        - _excel_row: número da linha real no Excel (para editar/excluir)
        Mais recentes primeiro.
        """
        self._ensure_workbook()
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]

        users: list[dict] = []

        for excel_row in range(2, ws.max_row + 1):  # pula cabeçalho
            values = [ws.cell(row=excel_row, column=c).value for c in range(1, len(HEADERS) + 1)]
            if all(v is None or str(v).strip() == "" for v in values):
                continue

            item = {HEADERS[i]: (values[i] if values[i] is not None else "") for i in range(len(HEADERS))}
            item["_excel_row"] = excel_row
            users.append(item)

        users.reverse()
        return users

    def update_user(self, excel_row: int, row: list):
        self._ensure_workbook()
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]

        if excel_row < 2 or excel_row > ws.max_row:
            raise ValueError("Linha inválida para atualização.")

        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=col_idx, value=value)

        wb.save(self.file_path)

    def delete_user(self, excel_row: int):
        self._ensure_workbook()
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]

        if excel_row < 2 or excel_row > ws.max_row:
            raise ValueError("Linha inválida para exclusão.")

        ws.delete_rows(excel_row, 1)
        wb.save(self.file_path)
